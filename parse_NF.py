import pandas as pd
from openpyxl import load_workbook
from unidecode import unidecode


FILE_PATH = "data/base/notasFiscais.xlsx"
SHEET_NAME = "Relatório"

def read_and_unmerge_excel(file_path, sheet_name):
  wb = load_workbook(file_path, data_only=True)
  ws = wb[sheet_name]
  
  merged_ranges = list(ws.merged_cells.ranges)
  for merged in merged_ranges:
    ws.unmerge_cells(str(merged))
    
  rows = []
  for row in ws.iter_rows(values_only=True):
    rows.append(list(row))
  return rows  

def normalize_header(header):
  new_header = []
  seen = {}
  
  for col in header:
    if col not in seen:
      seen[col] = 1
      new_header.append(col)
    else:
      seen[col] += 1
      new_header.append(f"{col}_{seen[col]}")
  
  return new_header

def extract_blocks(rows):
  data = []
  current_date = None
  current_cc = None
  header = None
  
  for row in rows:
    
    if (row and isinstance(row[0],str) and "Centro de custo" in row[0]):
      cc_value = row[3]
      if (isinstance(cc_value, str) and "obra" in cc_value.lower()):
        current_cc = cc_value
      
    if row and row[0] == "Data do movimento":
      current_date  = pd.to_datetime(row[3], dayfirst=True)
  
    if row and row[0] == "Movimento":
      header = normalize_header(row)
      continue
    
    if (header and row and isinstance(row[0], str) and row[0].startswith("NF")) and current_cc:
      record = dict(zip(header, row))
      record["data_movimento"] = current_date
      record["centro_custo"] = current_cc
      data.append(record)
      
  return pd.DataFrame(data)
      
def clean_df(df):
  df.columns = (df.columns.astype(str).str.strip().str.lower().map(unidecode).str.replace(" ","_").str.replace(".","", regex=False))
  return df

def rename_columns(df):
  return df.rename(columns={
    "forn": "fornecedor",
    "histórico_da_operação": "historico_operacao",
    "quantidade": "quantidade",
    "quantidade_2": "quantidade_unidade_basica",

    "unid": "unidade",
    "unid_2": "unidade_basica",

    "preço_unitário": "preco_unitario",
    "preço_unitário_2": "preco_unitario_unidade_basica",

    "total": "total",
    "total_2": "total_unidade_basica",          
  })
  
def clean_units(df):
    for col in ["unidade", "unidade_basica"]:
        if col in df.columns:
            df[col] = (
                df[col]
                .astype(str)
                .str.strip()
                .str.lower()
            )
    return df
  
def split_insumo(df):
  if "insumo" not in df.columns:
    return df
  
  split_data = df["insumo"].astype(str).str.split(" - ",n=1, expand=True)
  df["codigo_insumo"] = split_data[0].str.strip()
  if split_data.shape[1] > 1:
    df["descricao_insumo"] = split_data[1].str.strip()
  else:
    df["descricao_insumo"] = None
  
  df = df.drop(columns=["insumo"])
    
  return df

def drop_empty_columns_df(df):
  return df.dropna(axis=1, how="all")

def create_regional(df):
  if "centro_custo" not in df.columns:
    return df
  split_data = df["centro_custo"].astype(str).str.extract(r"^[^-]+-([^-]+)-")
  df["regional"] = split_data[0].str.strip()
  df.loc[df["regional"].str.lower().str.contains("mangabeiras|vista aruana", na=False), "regional"] = "SE"
  return df

def split_movimento(df):
    if "movimento" not in df.columns:
        return df

    df["movimento"] = (df["movimento"].astype(str).str.split("-", n=1).str[0])

    return df


def parse_excel_movimentos():
  rows = read_and_unmerge_excel(FILE_PATH, SHEET_NAME)
  df = extract_blocks(rows)
  df = drop_empty_columns_df(df)
  df = clean_df(df)
  df = rename_columns(df)
  df = clean_units(df)
  df = split_insumo(df)
  df = create_regional(df)
  df = split_movimento(df)
  
  return df 

if __name__ == "__main__":
  df = parse_excel_movimentos()
  nome_arquivo = 'data/1-data_frame_NF.xlsx'
  df.to_excel(nome_arquivo, index=False)
  print(f"Planilha '{nome_arquivo}' criada com sucesso!")
  

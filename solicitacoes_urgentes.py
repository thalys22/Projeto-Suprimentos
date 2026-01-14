import pandas as pd
from openpyxl import load_workbook
from parse_NF import clean_df
from parse_NFSE import split_fornecedor

FILE_PATH = "data/base/solicitacoesUrgentes.xlsx"
SHEET_NAME = "Sheet1"

def read_cells(file_path, sheet_name):
  wb = load_workbook(file_path, data_only=True)
  ws = wb[sheet_name]
  rows = []
  for row in ws.iter_rows(values_only=True):
    rows.append(list(row))
  return rows

def extract_blocks(rows):
    data = []
    header = None

    for row in rows:
        if row and isinstance(row[0], str) and row[0] == "Nº da Solicitação":
            header = row
            continue

        if header and row and isinstance(row[0], int):
            record = dict(zip(header, row))
            data.append(record)

    return pd.DataFrame(data)


def converter_data(df):
    df["Data do pedido"] = pd.to_datetime(
        df["Data do pedido"],
        dayfirst=True,
        errors="coerce"
    )

    df["Data da solicitação"] = pd.to_datetime(
        df["Data da solicitação"],
        dayfirst=True,
        errors="coerce"
    )
    
    df["Previsão de entrega"] = pd.to_datetime(
        df["Previsão de entrega"],
        dayfirst=True,
        errors="coerce"
    )
    
    df["Data entrega na obra"] = pd.to_datetime(
        df["Data entrega na obra"],
        dayfirst=True,
        errors="coerce"
    )
    

    return df


def insert_urgencia(df):
    df["dias"] = (
        df["Previsão de entrega"] - df["Data da solicitação"]
    ).dt.days

    df["urgencia"] = "Não urgente"

    df.loc[df["dias"] < 25, "urgencia"] = "Urgente"

    return df
  
def split_insumo(df):
    if "Descrição do insumo" not in df.columns:
        return df

    split_data = (
        df["Descrição do insumo"]
        .astype(str)
        .str.split(" - ", n=1, expand=True)
    )

    df["codigo_insumo"] = split_data[0].str.strip()

    if split_data.shape[1] > 1:
        df["descricao_insumo"] = split_data[1].str.strip()
    else:
        df["descricao_insumo"] = None

    df = df.drop(columns=["Descrição do insumo"])
    return df


    
def start():
    rows = read_cells(FILE_PATH, SHEET_NAME)
    df = extract_blocks(rows)

    df = df[df["Obra"].astype(str).str.lower().str.contains("obra", na=False)]
    df = converter_data(df)
    df = insert_urgencia(df)
    df = split_insumo(df)
    
    COLUNAS = [
    "Nº da Solicitação",
    "Obra",
    "codigo_insumo",
    "descricao_insumo",
    "Data da solicitação",
    "N° do Pedido",
    "Data do pedido",
    "Comprador",
    "Cód. Fornecedor",
    "Previsão de entrega",
    "Data entrega na obra",
    "N° da Nota fiscal",
    "Unidade de movimento",
    "urgencia"
]
    df = df[COLUNAS]

    df = df.rename(columns={
        "Nº da Solicitação": "numero_solicitacao",
        "insumo": "descricao_insumo",
        "Data da solicitação": "data_solicitacao",
        "Previsão de entrega": "previsao_de_entrega",
        "Cód. Fornecedor": "codigo_fornecedor",
        "Data do pedido": "data_pedido",
        "Data entrega na obra": "data_entrega_na_obra"
    })

    return df


if __name__ == "__main__":
  df = start()
  nome_arquivo = 'data/6-urgencias.xlsx'
  df.to_excel(nome_arquivo, index=False)
  print(f"Planilha '{nome_arquivo}' criada com sucesso!")
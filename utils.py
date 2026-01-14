import pandas as pd
from openpyxl import load_workbook
from unidecode import unidecode

def read_and_unmerge_excel(file_path, sheet_name):
    """Lê um arquivo Excel e desagrupa todas as células mescladas."""
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    
    merged_ranges = list(ws.merged_cells.ranges)
    for merged in merged_ranges:
        ws.unmerge_cells(str(merged))
        
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows

def normalize_header(header):
    """Normaliza o cabeçalho tratando colunas duplicadas."""
    new_header = []
    seen = {}
    
    for col in header:
        if col not in seen:
            seen[col] = 1
            new_header.append(col)
        else:
            seen[col] += 1
            new_header.append(f"{col}_{seen[col]}")
    
    return new_header

def clean_df(df):
    """Normaliza os nomes das colunas do DataFrame (lowercase, sem acentos, snake_case)."""
    df.columns = (
        df.columns.astype(str)
        .str.strip()
        .str.lower()
        .map(unidecode)
        .str.replace(" ", "_")
        .str.replace(".", "", regex=False)
    )
    return df

def drop_empty_columns_df(df):
    """Remove colunas que estão completamente vazias."""
    return df.dropna(axis=1, how="all")

def create_regional(df):
    """Extrai a regional do centro de custo e normaliza casos específicos."""
    if "centro_custo" not in df.columns:
        return df
    
    split_data = df["centro_custo"].astype(str).str.extract(r"^[^-]+-([^-]+)-")
    df["regional"] = split_data[0].str.strip()
    
    # Normalização de regionais específicas
    mask_se = df["regional"].str.lower().str.contains("mangabeiras|vista aruana", na=False)
    df.loc[mask_se, "regional"] = "SE"
    
    return df

def split_insumo(df, column_name="insumo"):
    """Separa o código e a descrição do insumo (formato 'Código - Descrição')."""
    if column_name not in df.columns:
        return df
    
    split_data = df[column_name].astype(str).str.split(" - ", n=1, expand=True)
    df["codigo_insumo"] = split_data[0].str.strip()
    
    if split_data.shape[1] > 1:
        df["descricao_insumo"] = split_data[1].str.strip()
    else:
        df["descricao_insumo"] = None
    
    df = df.drop(columns=[column_name])
    return df

def del_columns(df):
    """Remove colunas que começam com 'None'."""
    return df.loc[:, ~df.columns.str.startswith("None")]
